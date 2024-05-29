import json
from abc import ABC, abstractmethod
from typing import List, Union
from openpyxl import load_workbook


class chat_interface(ABC):

    @abstractmethod
    def request_chat(self, query: str, config: json) -> Union[dict, json]:
        pass

    # 针对4选1的xlsx
    def deal_one_out_of_four_xlsx(self, path: str, config: json) -> List[dict]:
        wb = load_workbook(filename=path)

        shs = wb.sheetnames

        res = []
        for sh in shs:
            ws = wb[sh]

            try:
                index = 1
                for row in ws.iter_rows(min_row=1, values_only=True):
                    # row 是一个元组，包含了当前行的所有数据
                    question = row[0]
                    options = [f'A:{row[1]}', f'B:{row[2]}', f'C:{row[3]}', f'D:{row[4]}']
                    answer = row[5]
                    question_type = row[6]

                    query = config["prompt"]["one_out_of_four"].format(question=question, options=options)
                    response = self.request_chat(query=query, config=config)
                    print(f'{index}.{query}')
                    print(response)
                    res.append(
                        {
                            "question": question,
                            "options": options,
                            "correct_answer": answer,
                            "type": question_type,
                            "kb_answer": response["answer"],
                        }
                    )
                    index += 1
            except Exception as e:
                print(f'迭代处理excel表格时发生异常:{e}')
                return res

        return res
