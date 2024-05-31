import json
from typing import List

import requests
from openpyxl import load_workbook

from evalute.utils.json2xlsx import is_xlsx


def deal_one_out_of_four_xlsx(path: str) -> List[dict]:
    wb = load_workbook(filename=path)

    shs = wb.sheetnames

    res = []
    for sh in shs:
        ws = wb[sh]

        try:
            index = 1
            for row in ws.iter_rows(min_row=1, values_only=True):
                # row 是一个元组，包含了当前行的所有数据
                question = row[0]
                print(f'{index}.{question}')
                options = [f'A:{row[1]}', f'B:{row[2]}', f'C:{row[3]}', f'D:{row[4]}']
                answer = row[5]
                question_type = row[6]

                response = request_kb_chat(question, options)
                print(response)
                res.append(
                    {
                        "question": question,
                        "options": options,
                        "correct_answer": answer,
                        "type": question_type,
                        "kb_answer": response["answer"],
                    }
                )
                index += 1
        except Exception as e:
            print(f'迭代处理excel表格时发生异常:{e}')
            return res

    return res


def request_kb_chat(question: str, options: List[str]) -> json:
    query = config["prompt"]["one_out_of_four"].format(question=question, options=options)
    print(query)
    args = config["chat_args"]

    request_body = {
        "query": query,
        "knowledge_base_name": args["knowledge_base_name"],
        "top_k": args["top_k"],
        "score_threshold": args["score_threshold"],
        "model_name": args["llm_models"][0],
        "temperature": args["temperature"],
        "prompt_name": args["prompt_name"][0],
    }

    try:
        response = requests.post(url=args["url"], headers={"Content-Type": "application/json"}, json=request_body)
    except Exception as e:
        response = '{\'answer\':\'发生异常' + str(e) + '\'}'

    return json.loads(response.text[response.text.find('{'):response.text.rfind('}') + 1])


if __name__ == '__main__':
    with open('./config/evaluate.json', 'r', encoding='utf-8') as cf:
        config = json.load(cf)

    srcs = config["sources"]
    targets = config["targets"]

    for i in range(len(srcs)):
        src = srcs[i]
        target = targets[i]

        if is_xlsx(src):
            dicts = deal_one_out_of_four_xlsx(src)
            with open(target, 'w', encoding='utf-8') as tf:
                json_str = json.dumps(dicts, ensure_ascii=False, indent=4)
                tf.write(json_str)
            print(f'{src}任务处理完毕,{target} 生成完成')
