import sys

from openpyxl import load_workbook
import random
from zhipuai import ZhipuAI
import re
from nltk.translate.bleu_score import sentence_bleu
from nltk.translate.bleu_score import SmoothingFunction
import jieba


# 计算bleu
def calculate_bleu(reference, hypothesis):
    reference_tokens = list(jieba.cut(reference))
    hypothesis_tokens = list(jieba.cut(hypothesis))
    smoothing_function = SmoothingFunction().method7
    score = sentence_bleu([reference_tokens], hypothesis_tokens, smoothing_function=smoothing_function)
    return score


# 将模型的回答转换成问题的列表
def extract_questions(text):
    # 使用正则表达式匹配以数字、空格、英文句号和冒号开头的行
    pattern = r'^\d+\s*[、.]?\s*问题\s*[:：]?\s*(.*)$'
    questions = []

    # 遍历文本的每一行
    for line in text.split('\n'):
        # 使用正则表达式进行匹配
        match = re.match(pattern, line)
        if match:
            question = match.group(1).strip()  # 提取匹配的问题，并去除两端的空格
            questions.append(question)

    return questions


if __name__ == '__main__':
    args = sys.argv

    origin_QA = args[1]
    key = args[2]
    client = ZhipuAI(api_key=key)  # 填写你自己的APIKey

    while True:
        file = open("./self_instruct/target/generation_instruction.txt", 'a', encoding="UTF-8")

        # 打开 Excel 文件
        workbook = load_workbook(origin_QA)
        # 获取工作表
        sheet = workbook.active
        # 遍历行和列
        input_list = []
        output_list = []
        for row in sheet.iter_rows(min_row=2, max_row=97, values_only=True):
            input_list.append(list(row)[0])
            output_list.append(list(row)[1])
        print(input_list)
        print(output_list)

        with open("./self_instruct/target/generation_instruction.txt", 'r', encoding="UTF-8") as file2:
            generation_instructions_list = file2.readlines()

        if len(generation_instructions_list) >= 5000:
            break

        # 随机从种子指令和生成指令中取5条
        seed_instructions = random.sample(input_list, 5)
        generation_instructions = random.sample(generation_instructions_list, 5)

        # 建立模版
        prompt = "你是一名精通周易的专家，下面是一些与周易相关的问题：\n"
        for index, instruction in enumerate(seed_instructions):
            prompt += f"{str(index + 1)}、问题：{instruction.strip()}\n"
        for index, instruction in enumerate(generation_instructions):
            prompt += f"{str(index + 6)}、问题：{instruction.strip()}\n"
        prompt += "请参考上面问题的内容和形式，对上述问题进行补充和拓展，再提出10个关于周易的比较深入的问题"
        print(prompt)

        global response
        # 向GLM4提问并返回答案
        try:
            response = client.chat.completions.create(
                model="glm-4",  # 填写需要调用的模型名称
                messages=[
                    {"role": "user",
                     "content": prompt}
                ],
            )
        except Exception as e:
            print(f"发生错误：{e}")
            continue

        print(response.choices[0].message.content)

        # 提取问题
        questions = extract_questions(response.choices[0].message.content)
        for i, question in enumerate(questions, start=1):
            print(f"{question}")

        # 指令存储到文件
        for question in questions:
            max_rouge = 0
            for generation_instruction in generation_instructions_list:
                bleu = calculate_bleu(question, generation_instruction)
                if bleu > max_rouge:
                    max_rouge = bleu
            if max_rouge < 0.7:
                print(f"录入：{question}")
                file.write(question + '\n')

        file.close()
        file2.close()
